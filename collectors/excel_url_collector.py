"""Excel URL 采集器。

下载公开 xlsx 文件，解析第一个或指定 sheet，把结果写入 raw_admission_records。
xls 暂不支持，返回 unsupported_xls。
"""

from pathlib import Path

from openpyxl import load_workbook

from collectors.common import load_field_mapping, load_parser_config, mark_duplicate_records
from collectors.file_download import download_file
from collectors.html_table_collector import (
    DEFAULT_HEADER_MAPPING,
    _append_sample,
    _build_headers_from_config,
    _build_message,
    _make_sample,
    _normalize_mapping_keys,
    _process_normal_row,
    _process_wide_row,
    _result,
    _safe_int_config,
    _safe_int_list_config,
    _safe_list_config,
    clean_cell_text,
    detect_header_row,
)
from db import create_connection


def _update_download_status(source_id: int, download_result: dict) -> None:
    conn = create_connection()
    try:
        conn.execute(
            """
            UPDATE raw_data_sources
            SET local_file_path = ?,
                file_size = ?,
                file_download_status = ?,
                file_download_message = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                download_result.get("local_file_path"),
                download_result.get("file_size"),
                download_result.get("status"),
                download_result.get("message"),
                source_id,
            ),
        )
        conn.commit()
    finally:
        conn.close()


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    return clean_cell_text(value)


def _load_excel_grid(local_file_path: str, config: dict) -> tuple[str, list[list[str]], str | None]:
    suffix = Path(local_file_path).suffix.lower()
    if suffix == ".xls":
        return "", [], "unsupported_xls：暂不支持 xls，请转换为 xlsx 或安装 xlrd 后扩展"
    if suffix != ".xlsx":
        return "", [], f"不支持的 Excel 文件类型：{suffix}"

    workbook = load_workbook(local_file_path, read_only=True, data_only=True)
    try:
        sheet_name = config.get("sheet_name")
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            sheet_index = _safe_int_config(config, "sheet_index", 0)
            if sheet_index >= len(workbook.sheetnames):
                sheet_index = 0
            worksheet = workbook[workbook.sheetnames[sheet_index]]

        grid = []
        for row in worksheet.iter_rows(values_only=True):
            grid.append([_cell_to_text(value) for value in row])
        return worksheet.title, grid, None
    finally:
        workbook.close()


def _parse_excel_grid(source: dict, grid: list[list[str]], sheet_name: str, preview: bool) -> dict:
    config = load_parser_config(source)
    fallback_header_row_index = _safe_int_config(config, "header_row_index", 0)
    skip_rows = _safe_int_config(config, "skip_rows", 0)
    fill_down_fields = _safe_list_config(config, "fill_down_fields")
    major_filter_keywords = _safe_list_config(config, "major_filter_keywords")
    header_keywords = _safe_list_config(config, "header_keywords")
    header_min_match_count = _safe_int_config(config, "header_min_match_count", 3)
    auto_detect_header = bool(config.get("auto_detect_header", False))
    auto_direction_tags = bool(config.get("auto_direction_tags", False))
    transform_type = config.get("transform_type")
    wide_subject_config = config.get("wide_subject_config", {})
    if not isinstance(wide_subject_config, dict):
        wide_subject_config = {}
    default_values = config.get("default_values", {})
    if not isinstance(default_values, dict):
        default_values = {}

    result = _result(source=source, table_index=0, preview=preview)
    result["parser_type"] = "excel_url"
    result["sheet_name"] = sheet_name

    custom_mapping = load_field_mapping(source)
    mapping = _normalize_mapping_keys(custom_mapping or DEFAULT_HEADER_MAPPING)

    if not grid:
        result["error_count"] = 1
        result["message"] = "Excel sheet 为空"
        return result

    auto_detected_header_row_index = None
    header_row_index = fallback_header_row_index
    if auto_detect_header:
        auto_detected_header_row_index = detect_header_row(
            grid=grid,
            header_keywords=header_keywords,
            min_match_count=header_min_match_count,
        )
        if auto_detected_header_row_index is not None:
            header_row_index = auto_detected_header_row_index

    if header_row_index >= len(grid):
        result["error_count"] = 1
        result["message"] = f"header_row_index={header_row_index} 超出范围"
        return result

    headers, effective_header_row_index = _build_headers_from_config(
        grid=grid,
        config=config,
        fallback_index=header_row_index,
    )
    header_row_index = effective_header_row_index
    result["header_row_index"] = header_row_index
    result["auto_detected_header_row_index"] = auto_detected_header_row_index
    result["detected_headers"] = headers

    header_row_indexes = _safe_int_list_config(config, "header_row_indexes")
    if header_row_indexes:
        data_start_index = max(header_row_indexes) + 1 + skip_rows
    else:
        data_start_index = header_row_index + 1 + skip_rows

    data_rows = grid[data_start_index:]
    result["total_data_rows"] = len(data_rows)
    last_fill_down_values: dict = {}

    conn = create_connection()
    try:
        for raw_cells in data_rows:
            if not any(clean_cell_text(cell) for cell in raw_cells):
                result["empty_row_skipped_count"] += 1
                result["skipped_count"] += 1
                _append_sample(result, _make_sample(raw_cells, {}, False, "empty_row"))
                continue

            try:
                if transform_type == "wide_subject_scores":
                    _process_wide_row(
                        conn=conn,
                        result=result,
                        source=source,
                        raw_cells=raw_cells,
                        headers=headers,
                        default_values=default_values,
                        auto_direction_tags=auto_direction_tags,
                        major_filter_keywords=major_filter_keywords,
                        preview=preview,
                        wide_config=wide_subject_config,
                    )
                else:
                    _process_normal_row(
                        conn=conn,
                        result=result,
                        source=source,
                        raw_cells=raw_cells,
                        headers=headers,
                        mapping=mapping,
                        fill_down_fields=fill_down_fields,
                        last_fill_down_values=last_fill_down_values,
                        default_values=default_values,
                        auto_direction_tags=auto_direction_tags,
                        major_filter_keywords=major_filter_keywords,
                        preview=preview,
                    )
            except Exception as exc:
                result["error_count"] += 1
                _append_sample(result, _make_sample(raw_cells, {}, False, f"error: {exc}"))

        if not preview:
            mark_duplicate_records(conn)
            conn.commit()
    finally:
        conn.close()

    if preview:
        result["inserted_count"] = 0

    result["message"] = _build_message(result)
    return result


def collect_excel_url_source(source: dict, preview: bool = False) -> dict:
    download_result = download_file(source["url"])
    _update_download_status(source["id"], download_result)

    if download_result.get("status") != "success":
        result = _result(source=source, table_index=0, preview=preview)
        result["parser_type"] = "excel_url"
        result["error_count"] = 1
        result["message"] = download_result.get("message") or "文件下载失败"
        return result

    sheet_name, grid, error_message = _load_excel_grid(
        download_result["local_file_path"],
        load_parser_config(source),
    )
    if error_message:
        result = _result(source=source, table_index=0, preview=preview)
        result["parser_type"] = "excel_url"
        result["sheet_name"] = sheet_name
        result["error_count"] = 1
        result["message"] = error_message
        return result

    return _parse_excel_grid(source, grid, sheet_name, preview)
