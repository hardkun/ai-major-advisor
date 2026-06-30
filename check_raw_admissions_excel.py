"""检查原始招生数据 Excel 是否适合导入 raw_admission_records。

运行命令：
    python check_raw_admissions_excel.py
"""

from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "data" / "raw_admissions_import_template.xlsx"


REQUIRED_FIELDS = [
    "raw_source_name",
    "school_name",
    "major_name",
    "admission_year",
    "admission_province",
    "subject_type",
    "min_score",
    "min_rank",
    "source_name",
]

INTEGER_FIELDS = [
    "admission_year",
    "min_score",
    "min_rank",
    "plan_count",
]

EXTRA_CHECK_FIELDS = [
    "direction_tags",
    "source_url",
]


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def can_parse_int(value) -> bool:
    try:
        int(value)
        return True
    except (TypeError, ValueError):
        return False


def split_direction_tags(value) -> list[str]:
    normalized = clean(value).replace("，", ",")
    return [tag.strip() for tag in normalized.split(",") if tag.strip()]


def add_error(errors: list[str], row_number: int | None, message: str) -> None:
    if row_number is None:
        errors.append(message)
    else:
        errors.append(f"第 {row_number} 行：{message}")


def add_warning(warnings: list[str], row_number: int | None, message: str) -> None:
    if row_number is None:
        warnings.append(message)
    else:
        warnings.append(f"第 {row_number} 行：{message}")


def read_excel_rows() -> tuple[list[str], list[dict]]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH, data_only=True)
    worksheet = workbook.worksheets[0]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [clean(value) for value in rows[0]]
    data_rows: list[dict] = []

    for excel_row_number, values in enumerate(rows[1:], start=2):
        if all(clean(value) == "" for value in values):
            continue
        row = {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}
        row["_excel_row_number"] = excel_row_number
        data_rows.append(row)

    return headers, data_rows


def check_row(
    row: dict,
    errors: list[str],
    warnings: list[str],
    school_names: set[str],
    major_names: set[str],
    direction_tag_set: set[str],
) -> None:
    row_number = row["_excel_row_number"]
    school_name = clean(row.get("school_name"))
    major_name = clean(row.get("major_name"))
    direction_tags = clean(row.get("direction_tags"))

    if school_name:
        school_names.add(school_name)
    if major_name:
        major_names.add(major_name)

    for tag in split_direction_tags(direction_tags):
        direction_tag_set.add(tag)

    for field in REQUIRED_FIELDS:
        if not clean(row.get(field)):
            add_error(errors, row_number, f"必填字段 {field} 为空")

    for field in INTEGER_FIELDS:
        value = row.get(field)
        if clean(value) and not can_parse_int(value):
            add_error(errors, row_number, f"数字字段 {field} 不是有效整数：{clean(value)}")

    min_rank = row.get("min_rank")
    if clean(min_rank) and can_parse_int(min_rank) and int(min_rank) <= 0:
        add_error(errors, row_number, f"min_rank 必须大于 0，当前值：{clean(min_rank)}")

    if not direction_tags:
        add_warning(warnings, row_number, "direction_tags 为空，可能影响后续方向匹配")

    if not clean(row.get("source_url")):
        add_warning(warnings, row_number, "source_url 为空，建议补充正式数据来源链接")


def check_total_rows(total_rows: int, errors: list[str], warnings: list[str]) -> None:
    if total_rows < 10:
        add_error(errors, None, "数据量少于 10 条，不建议导入")
    elif total_rows < 30:
        add_warning(warnings, None, "当前数据少于 30 条，只适合功能测试")


def print_messages(title: str, messages: list[str]) -> None:
    if not messages:
        return

    print(f"\n{title}：")
    for message in messages:
        print(f"- {message}")


def main() -> None:
    errors: list[str] = []
    warnings: list[str] = []
    school_names: set[str] = set()
    major_names: set[str] = set()
    direction_tag_set: set[str] = set()

    try:
        headers, rows = read_excel_rows()
    except FileNotFoundError as exc:
        print(exc)
        return

    if not headers:
        print("Excel 文件为空或缺少表头。")
        return

    fields_to_check = REQUIRED_FIELDS + INTEGER_FIELDS + EXTRA_CHECK_FIELDS
    missing_columns = [field for field in fields_to_check if field not in headers]
    for field in missing_columns:
        add_error(errors, None, f"表头缺少字段：{field}")

    for row in rows:
        check_row(
            row=row,
            errors=errors,
            warnings=warnings,
            school_names=school_names,
            major_names=major_names,
            direction_tag_set=direction_tag_set,
        )

    total_rows = len(rows)
    check_total_rows(total_rows, errors, warnings)

    print("原始招生数据 Excel 检查结果")
    print(f"文件路径：{EXCEL_PATH}")
    print(f"总行数：{total_rows}")
    print(f"error 数量：{len(errors)}")
    print(f"warning 数量：{len(warnings)}")
    print(f"涉及院校数量：{len(school_names)}")
    print(f"涉及专业数量：{len(major_names)}")

    if direction_tag_set:
        sorted_tags = sorted(direction_tag_set)
        print(f"涉及方向标签集合：{', '.join(sorted_tags)}")
    else:
        print("涉及方向标签集合：无")

    print_messages("Errors", errors)
    print_messages("Warnings", warnings)

    print()
    if errors:
        print("检查结论：不建议导入，请先修正错误。")
    elif warnings:
        print("检查结论：可以导入，但仅适合测试或小范围验证。")
    else:
        print("检查结论：数据检查通过，可以导入。")


if __name__ == "__main__":
    main()
