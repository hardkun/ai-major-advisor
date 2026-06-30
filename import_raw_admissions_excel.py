"""将 Excel 批量导入 raw_admission_records。

运行命令：
    python import_raw_admissions_excel.py

重要说明：
    本脚本只把 Excel 数据写入 raw_admission_records，状态默认为 pending。
    它不会直接写入 admissions，也不会影响推荐接口。
"""

from pathlib import Path

from openpyxl import load_workbook

from crud.raw_admission_records import mark_duplicate_records
from db import create_connection, init_db


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "data" / "raw_admissions_import_template.xlsx"


RAW_RECORD_FIELDS = [
    "raw_source_id",
    "school_name",
    "school_code",
    "school_province",
    "city",
    "school_level",
    "school_tags",
    "major_name",
    "major_code",
    "major_category",
    "direction_tags",
    "major_description",
    "career_paths",
    "admission_year",
    "admission_province",
    "subject_type",
    "batch",
    "major_group_code",
    "elective_requirement",
    "min_score",
    "min_rank",
    "plan_count",
    "tuition",
    "campus",
    "source_name",
    "source_url",
    "raw_text",
    "status",
]


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def empty_to_none(value) -> str | None:
    cleaned = clean(value)
    return cleaned or None


def parse_optional_int(value, field_name: str, row_number: int) -> int | None:
    cleaned = empty_to_none(value)
    if cleaned is None:
        return None
    try:
        return int(cleaned)
    except ValueError as exc:
        raise ValueError(
            f"第 {row_number} 行的 {field_name} 不是有效整数：{cleaned}"
        ) from exc


def read_excel_rows() -> list[dict]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [clean(value) for value in rows[0]]
    result: list[dict] = []
    for excel_row_number, values in enumerate(rows[1:], start=2):
        if all(clean(value) == "" for value in values):
            continue
        row = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
        }
        row["_excel_row_number"] = excel_row_number
        result.append(row)

    return result


def get_or_create_raw_data_source(db, row: dict) -> tuple[int, bool]:
    raw_source_name = empty_to_none(row.get("raw_source_name"))
    if raw_source_name is None:
        raise ValueError("raw_source_name 不能为空")

    existing = db.execute(
        """
        SELECT id FROM raw_data_sources
        WHERE name = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (raw_source_name,),
    ).fetchone()
    if existing:
        return existing["id"], False

    cursor = db.execute(
        """
        INSERT INTO raw_data_sources
            (name, source_type, url, parser_type, enabled, description)
        VALUES (?, ?, ?, ?, 1, ?)
        """,
        (
            raw_source_name,
            empty_to_none(row.get("raw_source_type")),
            empty_to_none(row.get("raw_source_url")),
            empty_to_none(row.get("parser_type")),
            "由 Excel 批量导入时自动创建",
        ),
    )
    return cursor.lastrowid, True


def raw_record_exists(db, data: dict) -> bool:
    row = db.execute(
        """
        SELECT id FROM raw_admission_records
        WHERE COALESCE(school_name, '') = COALESCE(?, '')
          AND COALESCE(major_name, '') = COALESCE(?, '')
          AND COALESCE(admission_year, -1) = COALESCE(?, -1)
          AND COALESCE(admission_province, '') = COALESCE(?, '')
          AND COALESCE(subject_type, '') = COALESCE(?, '')
          AND COALESCE(major_group_code, '') = COALESCE(?, '')
          AND COALESCE(major_code, '') = COALESCE(?, '')
          AND COALESCE(min_score, -1) = COALESCE(?, -1)
          AND COALESCE(min_rank, -1) = COALESCE(?, -1)
        LIMIT 1
        """,
        (
            data["school_name"],
            data["major_name"],
            data["admission_year"],
            data["admission_province"],
            data["subject_type"],
            data["major_group_code"],
            data["major_code"],
            data["min_score"],
            data["min_rank"],
        ),
    ).fetchone()
    return row is not None


def build_raw_record_data(row: dict) -> dict:
    row_number = row["_excel_row_number"]
    return {
        "raw_source_id": row["raw_source_id"],
        "school_name": empty_to_none(row.get("school_name")),
        "school_code": empty_to_none(row.get("school_code")),
        "school_province": empty_to_none(row.get("school_province")),
        "city": empty_to_none(row.get("city")),
        "school_level": empty_to_none(row.get("school_level")),
        "school_tags": empty_to_none(row.get("school_tags")),
        "major_name": empty_to_none(row.get("major_name")),
        "major_code": empty_to_none(row.get("major_code")),
        "major_category": empty_to_none(row.get("major_category")),
        "direction_tags": empty_to_none(row.get("direction_tags")),
        "major_description": empty_to_none(row.get("major_description")),
        "career_paths": empty_to_none(row.get("career_paths")),
        "admission_year": parse_optional_int(
            row.get("admission_year"),
            "admission_year",
            row_number,
        ),
        "admission_province": empty_to_none(row.get("admission_province")),
        "subject_type": empty_to_none(row.get("subject_type")),
        "batch": empty_to_none(row.get("batch")),
        "major_group_code": empty_to_none(row.get("major_group_code")),
        "elective_requirement": empty_to_none(row.get("elective_requirement")),
        "min_score": parse_optional_int(row.get("min_score"), "min_score", row_number),
        "min_rank": parse_optional_int(row.get("min_rank"), "min_rank", row_number),
        "plan_count": parse_optional_int(
            row.get("plan_count"),
            "plan_count",
            row_number,
        ),
        "tuition": empty_to_none(row.get("tuition")),
        "campus": empty_to_none(row.get("campus")),
        "source_name": empty_to_none(row.get("source_name")),
        "source_url": empty_to_none(row.get("source_url")),
        "raw_text": empty_to_none(row.get("raw_text")),
        "status": "pending",
    }


def insert_raw_record(db, data: dict) -> int:
    columns = ", ".join(RAW_RECORD_FIELDS)
    placeholders = ", ".join(["?"] * len(RAW_RECORD_FIELDS))
    values = [data[field] for field in RAW_RECORD_FIELDS]

    cursor = db.execute(
        f"""
        INSERT INTO raw_admission_records ({columns})
        VALUES ({placeholders})
        """,
        values,
    )
    return cursor.lastrowid


def main() -> None:
    init_db()
    rows = read_excel_rows()

    created_raw_sources = 0
    created_raw_records = 0
    skipped_duplicates = 0
    error_rows = 0
    error_messages: list[str] = []

    db = create_connection()
    try:
        for row in rows:
            row_number = row["_excel_row_number"]
            try:
                raw_source_id, created_source = get_or_create_raw_data_source(db, row)
                if created_source:
                    created_raw_sources += 1

                row["raw_source_id"] = raw_source_id
                data = build_raw_record_data(row)
                if raw_record_exists(db, data):
                    skipped_duplicates += 1
                    continue

                insert_raw_record(db, data)
                created_raw_records += 1
            except Exception as exc:
                error_rows += 1
                error_messages.append(f"第 {row_number} 行导入失败：{exc}")

        db.commit()
        mark_duplicate_records(db)
    finally:
        db.close()

    print("原始招生数据 Excel 导入完成")
    print(f"新增 raw_data_sources 数量：{created_raw_sources}")
    print(f"新增 raw_admission_records 数量：{created_raw_records}")
    print(f"跳过重复数据数量：{skipped_duplicates}")
    print(f"错误行数量：{error_rows}")

    if error_messages:
        print("\n错误明细：")
        for message in error_messages:
            print(f"- {message}")


if __name__ == "__main__":
    main()
